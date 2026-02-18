#!/usr/bin/env python3
"""Simple password generator that writes accepted password to Excel and optionally password-protects it (Windows/Excel).

Usage: place a dictionary.txt file in the same folder (one word per line), then run this script.
"""
import os
import random
import string
import sys

from pathlib import Path
try:
    from colorama import init as colorama_init, Fore, Style
    colorama_init(autoreset=True)
except Exception:
    # colorama is optional; output will still work without colors
    class Fore:
        RED = ''
        GREEN = ''
        YELLOW = ''
        CYAN = ''
        MAGENTA = ''

    class Style:
        BRIGHT = ''
        RESET_ALL = ''

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except Exception as e:
    print("Missing dependency openpyxl. Install with: pip install -r requirements.txt")
    raise

SYMBOLS = list('!"£$%@#*^')
EXCEL_DIR = "TEMP EXCEL PASSWORDS"
EMAIL_DIR = "READY EMAILS"
# Ensure output folders exist
Path(EXCEL_DIR).mkdir(parents=True, exist_ok=True)
Path(EMAIL_DIR).mkdir(parents=True, exist_ok=True)

def load_dictionary(path: str = "dictionary.txt"):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Dictionary file not found: {path}")
    words = [w.strip() for w in p.read_text(encoding="utf-8", errors="ignore").splitlines()]
    words = [w for w in words if w]
    if not words:
        raise ValueError("Dictionary file is empty")
    return words


def build_candidate_intact(w1: str, w2: str):
    # Keep words intact and readable: Title case and remove surrounding whitespace
    a = ''.join(ch for ch in w1.strip() if ch.isalpha()).capitalize()
    b = ''.join(ch for ch in w2.strip() if ch.isalpha()).capitalize()
    if not a:
        a = ''.join(ch for ch in w1.strip())[:3].capitalize()
    if not b:
        b = ''.join(ch for ch in w2.strip())[:3].capitalize()
    return a, b


def ensure_requirements_readable(password: str):
    # Ensure password has at least one digit and at least one symbol (but no more than two symbols).
    s = password
    if not any(c.isdigit() for c in s):
        s = s + random.choice(string.digits)
    present_symbols = [c for c in s if c in SYMBOLS]
    if not present_symbols:
        s = s + random.choice(SYMBOLS)
    return s


def generate_password(words, max_attempts=1000):
    patterns = [
        # Each entry is a tuple: (callable, symbol_slots)
        # - callable: function(a, b, d, s1, s2) that builds a password string
        # - symbol_slots: 1 or 2 indicating how many distinct symbol placeholders
        #   the callable expects (s2 may be empty for a single-slot pattern)
        (lambda a, b, d, s1, s2: f"{a}{d}{b}{s1}", 1),
        (lambda a, b, d, s1, s2: f"{d}{a}{s1}{b}{s2}", 2),
        (lambda a, b, d, s1, s2: f"{a}{s1}{b}{d}", 1),
        (lambda a, b, d, s1, s2: f"{a}{s1}{b}{s2}{d}", 2),
        (lambda a, b, d, s1, s2: f"{d}{a}{b}{s1}", 1),
    ]

    for _ in range(max_attempts):
        w1 = random.choice(words)
        w2 = random.choice(words)
        a, b = build_candidate_intact(w1, w2)
        d = random.choice(string.digits)
        pattern, slots = random.choice(patterns)
        if slots == 1:
            s1 = random.choice(SYMBOLS)
            s2 = ''
        else:
            s1, s2 = random.sample(SYMBOLS, 2)
        pwd = pattern(a, b, d, s1, s2)
        # Ensure password length is within bounds. If shorter than minimum (10),
        # add padding while preserving the intact word parts:
        # - If the password contains no symbol, add exactly one unique symbol to satisfy
        #   the requirement of at least one symbol (but never add more than two symbols).
        # - Fill remaining padding with digits only to avoid introducing additional symbols.
        if len(pwd) < 10:
            need = 10 - len(pwd)
            present_symbols = [c for c in pwd if c in SYMBOLS]
            padding_chars = []
            # If no symbol present, ensure we add exactly one unique symbol (to meet min-1)
            if not present_symbols and need > 0:
                # pick one symbol not already present
                sym = random.choice([x for x in SYMBOLS if x not in present_symbols])
                padding_chars.append(sym)
                need -= 1
            # fill remaining padding with digits only to avoid exceeding 2 symbols
            if need > 0:
                padding_chars.extend(random.choices(string.digits, k=need))
            random.shuffle(padding_chars)
            pwd = pwd + ''.join(padding_chars)
        if 10 <= len(pwd) <= 32:
            pwd = ensure_requirements_readable(pwd)
            if 10 <= len(pwd) <= 32:
                return pwd
        # if too long, try another pair
        continue
    raise RuntimeError("Failed to generate password within constraints")


def write_password_to_excel(password: str, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Password"
    ws["A1"] = password
    ws["A1"].font = Font(size=36)
    # approximate column width and row height to fit font
    width = max(10, int(len(password) * 1.2))
    ws.column_dimensions['A'].width = width
    ws.row_dimensions[1].height = int(36 * 1.3)
    wb.save(out_path)


def protect_excel_with_password(path: str, password: str):
    # Use pywin32 COM automation to password-protect the workbook (requires Excel on Windows)
    try:
        import time
        from win32com.client import Dispatch

        abs_path = os.path.abspath(path)
        excel = Dispatch('Excel.Application')
        excel.DisplayAlerts = False
        excel.Visible = False
        wb = excel.Workbooks.Open(abs_path)
        # Auto-fit the first column and first row for the password cell
        try:
            ws = wb.Worksheets(1)
            # AutoFit column A and row 1
            ws.Columns('A').AutoFit()
            ws.Rows(1).AutoFit()
        except Exception:
            pass
        # SaveAs with FileFormat=51 (xlsx) and Password parameter sets the workbook open password
        # VBA signature: SaveAs(Filename, FileFormat, Password, ...)
        wb.SaveAs(abs_path, 51, password)
        wb.Close(SaveChanges=True)
        excel.Quit()
        # allow COM to clean up
        time.sleep(0.5)
        return True
    except Exception as e:
        print("Warning: Could not apply Excel password protection via COM. File saved unprotected.")
        print(f"Details: {e}")
        return False


def attach_xlsx_to_eml(xlsx_path: str, templates_dir: str = "EXTERNAL EMAIL TEMPLATES", ready_dir: str = EMAIL_DIR):
    """Present available .eml templates and attach the given Excel file to the chosen template.

    Returns the path to the saved .eml file or None if cancelled / not possible.
    """
    from pathlib import Path
    from email import message_from_binary_file
    from email.policy import default

    tpl_dir = Path(templates_dir)
    if not tpl_dir.exists() or not tpl_dir.is_dir():
        print(f"Templates folder not found: '{templates_dir}'")
        return None

    eml_files = sorted([p for p in tpl_dir.glob("*.eml") if p.is_file()])
    if not eml_files:
        print(f"No .eml templates found in '{templates_dir}'.")
        return None

    print("Select a template to attach the Excel file to:")
    for i, p in enumerate(eml_files, start=1):
        print(f"{i}: {p.name}")
    print("0: Cancel")

    while True:
        sel = input(f"Enter number (0-{len(eml_files)}): ").strip()
        if sel == '0':
            print("Attachment cancelled.")
            return None
        if not sel.isdigit():
            print("Please enter a number.")
            continue
        idx = int(sel) - 1
        if 0 <= idx < len(eml_files):
            chosen = eml_files[idx]
            break
        print("Invalid selection, try again.")

    # Load template message
    try:
        with chosen.open('rb') as f:
            msg = message_from_binary_file(f, policy=default)
    except Exception as e:
        print(f"Failed to load template '{chosen}': {e}")
        return None

    # Prompt for recipient email (optional) and set/replace the To header
    try:
        recipient = input("Enter recipient email address (paste and press Enter). Leave blank to keep template recipient: ").strip()
        if recipient:
            try:
                if 'To' in msg:
                    msg.replace_header('To', recipient)
                else:
                    msg['To'] = recipient
            except Exception:
                # Fallback: set header directly
                msg['To'] = recipient
    except Exception:
        # Don't fail attachment if input can't be read for any reason
        recipient = ''

    # Read excel bytes
    xlsx_path_obj = Path(xlsx_path)
    if not xlsx_path_obj.exists():
        print(f"Excel file not found: {xlsx_path}")
        return None
    try:
        with xlsx_path_obj.open('rb') as f:
            xlsx_data = f.read()
    except Exception as e:
        print(f"Failed to read Excel file: {e}")
        return None

    xlsx_name = xlsx_path_obj.name

    # Attach the workbook (email lib will convert to multipart if needed)
    try:
        msg.add_attachment(
            xlsx_data,
            maintype='application',
            subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=xlsx_name,
        )
    except Exception as e:
        print(f"Failed to attach file to message: {e}")
        return None

    # Save output .eml into the ready emails directory
    out_filename = f"READY_TO_SEND_{chosen.stem}_{xlsx_name}.eml"
    out_dir = Path(ready_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / out_filename
    
    # Check if file exists and handle accordingly
    if out_path.exists():
        print(f"File '{out_filename}' already exists in '{ready_dir}'.")
        print("Choose an action: (E)nter new filename  (O)verwrite file  (C)ancel")
        choice = input("Enter choice [E/O/C] (default E): ").strip().lower() or 'e'
        if choice.startswith('e'):
            # Generate new filename with counter
            counter = 1
            base_name = out_filename.replace('.eml', '')
            while True:
                out_filename = f"{base_name}_{counter}.eml"
                out_path = out_dir / out_filename
                if not out_path.exists():
                    break
                counter += 1
        elif choice.startswith('o'):
            pass  # Proceed with overwrite
        elif choice.startswith('c'):
            print("Email attachment cancelled.")
            return None
        else:
            # Invalid choice, default to rename
            counter = 1
            base_name = out_filename.replace('.eml', '')
            while True:
                out_filename = f"{base_name}_{counter}.eml"
                out_path = out_dir / out_filename
                if not out_path.exists():
                    break
                counter += 1
    
    try:
        with out_path.open('wb') as f:
            f.write(msg.as_bytes(policy=default))
    except Exception as e:
        print(f"Failed to save attached email: {e}")
        return None

    print(Fore.YELLOW + f"Saved attached email as: {out_path}" + Style.RESET_ALL)
    return out_path


def prompt_yes_no(prompt: str) -> bool:
    while True:
        ans = input(Fore.CYAN + prompt + " [y/n]: " + Style.RESET_ALL).strip().lower()
        if ans in ('y', 'yes'):
            return True
        if ans in ('n', 'no'):
            return False


def main():
    try:
        words = load_dictionary('dictionary.txt')
    except Exception as e:
        print(e)
        sys.exit(1)

    while True:
        pwd = generate_password(words)
        print('\n' + Fore.MAGENTA + Style.BRIGHT + '=== Generated Password ===' + Style.RESET_ALL)
        print('\n' + Fore.GREEN + Style.BRIGHT + pwd + Style.RESET_ALL + '\n')
        # info panel
        length = len(pwd)
        has_upper = any(c.isupper() for c in pwd)
        has_lower = any(c.islower() for c in pwd)
        has_digit = any(c.isdigit() for c in pwd)
        has_symbol = any(c in SYMBOLS for c in pwd)
        print(Fore.YELLOW + f"Length: {length}    Upper: {has_upper}    Lower: {has_lower}    Digit: {has_digit}    Symbol: {has_symbol}" + Style.RESET_ALL)
        print(Fore.MAGENTA + '===========================' + Style.RESET_ALL)
        resp = prompt_yes_no('Accept this password?')
        print()
        if resp:
            break
    # write to excel with safe overwrite handling (store in TEMP EXCEL PASSWORDS)
    default_name = 'password.xlsx'
    while True:
        base_name = input(f"Enter output Excel filename [{default_name}]: ").strip() or default_name
        if not base_name.lower().endswith('.xlsx'):
            base_name += '.xlsx'
        out_path = Path(EXCEL_DIR) / base_name
        if out_path.exists():
            print(f"File '{out_path.name}' already exists in '{EXCEL_DIR}'.")
            print("Choose an action: (E)nter new filename  (O)verwrite file  (C)ancel")
            choice = input("Enter choice [E/O/C] (default E): ").strip().lower() or 'e'
            if choice.startswith('e'):
                continue
            if choice.startswith('o'):
                break
            if choice.startswith('c'):
                print('Operation cancelled.')
                sys.exit(0)
            continue
        else:
            break

    try:
        write_password_to_excel(pwd, str(out_path))
        print('\n' + Fore.CYAN + f"Password written to first cell of {out_path}" + Style.RESET_ALL)
    except Exception as e:
        print(Fore.RED + f"Failed to write Excel file: {e}" + Style.RESET_ALL)
        sys.exit(1)

    print()
    protect = prompt_yes_no('Would you like to password-protect the Excel file (requires Excel on Windows) with an external weekly password?')
    print()
    if protect:
        weekly = input('Paste external weekly password (will be used to protect Excel file): ').rstrip('\n')
        if weekly:
            ok = protect_excel_with_password(str(out_path), weekly)
            if ok:
                print('Excel file protected with provided password.')
            else:
                print('Excel file left unprotected (see warning above).')
        else:
            print('No password entered; skipping protection.')

    # Offer to attach the generated Excel file to an .eml template
    try:
        print()
        attach_resp = prompt_yes_no('Attach the generated Excel file to an .eml template from EXTERNAL EMAIL TEMPLATES?')
        print()
        if attach_resp:
            attach_xlsx_to_eml(str(out_path), ready_dir=EMAIL_DIR)
    except Exception:
        # Defensive: do not let attachment errors break existing workflow
        print('Attachment step skipped due to unexpected error.')

    # Workflow reminders: confirm AD and M365 password set manually
    try:
        print()
        ad_resp = prompt_yes_no('Confirm password set for Active Directory?')
        print()
        if not ad_resp:
            print(Fore.YELLOW + 'Reminder: Set Active Directory password manually before finishing.' + Style.RESET_ALL)
    except Exception:
        pass

    try:
        m365_resp = prompt_yes_no('Confirm password set for M365?')
        print()
        if not m365_resp:
            print(Fore.YELLOW + 'Reminder: Set M365 password manually before finishing.' + Style.RESET_ALL)
    except Exception:
        pass

    # Attempt to open the READY EMAILS folder in File Explorer (Windows)
    try:
        folder_path = os.path.abspath(EMAIL_DIR)
        if os.path.isdir(folder_path):
            try:
                os.startfile(folder_path)
                print(Fore.CYAN + f"Opened ready emails folder: {folder_path}" + Style.RESET_ALL)
            except Exception:
                # Fallback to explorer via subprocess
                try:
                    import subprocess
                    subprocess.run(['explorer', folder_path])
                    print(Fore.CYAN + f"Opened ready emails folder via explorer: {folder_path}" + Style.RESET_ALL)
                except Exception as e:
                    print(Fore.YELLOW + f"Could not open READY EMAILS folder: {e}" + Style.RESET_ALL)
        else:
            print(Fore.YELLOW + f"Ready emails folder not found: {folder_path}" + Style.RESET_ALL)
    except Exception as e:
        print(Fore.YELLOW + f"Could not open READY EMAILS folder: {e}" + Style.RESET_ALL)

    print('Done.')


if __name__ == '__main__':
    main()
